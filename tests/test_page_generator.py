from __future__ import annotations

import json

import pytest

from sre_web_inspector.config_schema import AppConfig, PageGeneratorConfig
from sre_web_inspector.page_generator import expand_page_generators

# Import main.py functions for integration coverage
from main import load_and_validate_config


class TestExpandIds:
    def test_simple_ids(self):
        config = {
            "page_generators": [
                {
                    "name": "test_gen",
                    "type": "ids",
                    "id_field": "id",
                    "ids": [1001, 1002, 1003],
                    "template": {
                        "name": "page_{{ id }}",
                        "url": "https://example.com/ids/{{ id }}/query",
                    },
                }
            ]
        }
        vars_map = {}
        pages = expand_page_generators(config, vars_map)
        assert len(pages) == 3
        assert pages[0]["name"] == "page_1001"
        assert pages[0]["url"] == "https://example.com/ids/1001/query"
        assert pages[0]["_generated"] is True
        assert pages[0]["_generator"] == "test_gen"
        assert pages[1]["name"] == "page_1002"
        assert pages[2]["name"] == "page_1003"

    def test_ids_with_vars(self):
        config = {
            "page_generators": [
                {
                    "name": "pod_gen",
                    "type": "ids",
                    "id_field": "pod",
                    "ids": ["pod-a", "pod-b"],
                    "template": {
                        "name": "pod_{{ pod }}",
                        "url": "{{ base_url }}/k8s/{{ namespace }}/pods/{{ pod }}",
                    },
                }
            ]
        }
        vars_map = {"base_url": "https://k8s.example.com", "namespace": "prod"}
        pages = expand_page_generators(config, vars_map)
        assert len(pages) == 2
        assert pages[0]["url"] == "https://k8s.example.com/k8s/prod/pods/pod-a"
        assert pages[1]["url"] == "https://k8s.example.com/k8s/prod/pods/pod-b"

    def test_ids_preserves_type(self):
        config = {
            "page_generators": [
                {
                    "name": "size_gen",
                    "type": "ids",
                    "id_field": "size",
                    "ids": [100, 200, 500],
                    "template": {
                        "name": "size_{{ size }}",
                        "url": "https://example.com/items",
                        "timeout": "{{ size }}",
                    },
                }
            ]
        }
        vars_map = {}
        pages = expand_page_generators(config, vars_map)
        assert len(pages) == 3
        # render_value preserves type for whole-string templates, so timeout should be int
        assert pages[0]["timeout"] == 100
        assert pages[2]["timeout"] == 500


class TestExpandList:
    def test_simple_list(self):
        config = {
            "page_generators": [
                {
                    "name": "multi",
                    "type": "list",
                    "values": [
                        {"name": "page_a", "url_path": "/a"},
                        {"name": "page_b", "url_path": "/b"},
                    ],
                    "template": {
                        "name": "{{ name }}",
                        "url": "https://example.com{{ url_path }}",
                    },
                }
            ]
        }
        vars_map = {}
        pages = expand_page_generators(config, vars_map)
        assert len(pages) == 2
        assert pages[0]["name"] == "page_a"
        assert pages[0]["url"] == "https://example.com/a"
        assert pages[1]["name"] == "page_b"
        assert pages[1]["url"] == "https://example.com/b"

    def test_list_with_extra_fields(self):
        config = {
            "page_generators": [
                {
                    "name": "resources",
                    "type": "list",
                    "values": [
                        {"id": "res1", "label": "Resource 1"},
                        {"id": "res2", "label": "Resource 2"},
                    ],
                    "template": {
                        "name": "resource_{{ id }}",
                        "url": "https://example.com/resources/{{ id }}",
                        "screenshot": True,
                        "save_html": False,
                    },
                }
            ]
        }
        vars_map = {}
        pages = expand_page_generators(config, vars_map)
        assert len(pages) == 2
        assert pages[0]["screenshot"] is True
        assert pages[0]["save_html"] is False

    def test_list_with_global_vars(self):
        config = {
            "page_generators": [
                {
                    "name": "env_pages",
                    "type": "list",
                    "values": [
                        {"env": "dev"},
                        {"env": "prod"},
                    ],
                    "template": {
                        "name": "{{ env }}_dashboard",
                        "url": "{{ base_url }}/{{ env }}/dashboard",
                    },
                }
            ]
        }
        vars_map = {"base_url": "https://monitor.example.com"}
        pages = expand_page_generators(config, vars_map)
        assert pages[0]["url"] == "https://monitor.example.com/dev/dashboard"
        assert pages[1]["url"] == "https://monitor.example.com/prod/dashboard"


class TestMaxPages:
    def test_within_limit(self):
        config = {
            "page_generators": [
                {
                    "name": "small",
                    "type": "ids",
                    "ids": [1, 2, 3],
                    "max_pages": 5,
                    "template": {"name": "p{{ id }}", "url": "https://x.com/{{ id }}"},
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 3

    def test_exceeds_limit(self):
        config = {
            "page_generators": [
                {
                    "name": "big",
                    "type": "ids",
                    "ids": [1, 2, 3, 4, 5],
                    "max_pages": 3,
                    "template": {"name": "p{{ id }}", "url": "https://x.com/{{ id }}"},
                }
            ]
        }
        with pytest.raises(ValueError, match="exceeding max_pages"):
            expand_page_generators(config, {})


class TestEmptyGenerators:
    def test_no_generators(self):
        config = {}
        pages = expand_page_generators(config, {})
        assert pages == []

    def test_none_generators(self):
        config = {"page_generators": None}
        pages = expand_page_generators(config, {})
        assert pages == []

    def test_empty_list_values(self):
        config = {
            "page_generators": [
                {"name": "empty", "type": "list", "values": [], "template": {"name": "x", "url": "https://x.com"}},
            ]
        }
        pages = expand_page_generators(config, {})
        assert pages == []

    def test_empty_ids(self):
        config = {
            "page_generators": [
                {"name": "empty_ids", "type": "ids", "ids": [], "template": {"name": "x", "url": "https://x.com"}},
            ]
        }
        pages = expand_page_generators(config, {})
        assert pages == []


class TestPageGeneratorSchema:
    def test_valid_list_generator(self):
        gen = PageGeneratorConfig(
            name="test",
            type="list",
            values=[{"id": "1"}],
            template={"name": "p", "url": "https://x.com"},
        )
        assert gen.name == "test"
        assert gen.type == "list"

    def test_valid_ids_generator(self):
        gen = PageGeneratorConfig(
            name="ids_test",
            type="ids",
            ids=[1, 2, 3],
            template={"name": "p{{ id }}", "url": "https://x.com/{{ id }}"},
        )
        assert len(gen.ids) == 3

    def test_template_required(self):
        with pytest.raises(Exception):
            PageGeneratorConfig(name="bad", type="list", values=[{"x": 1}])

    def test_default_max_pages(self):
        gen = PageGeneratorConfig(
            name="test", type="list", values=[], template={"name": "p", "url": "https://x.com"},
        )
        assert gen.max_pages == 500


class TestAppConfigIntegration:
    def test_app_config_with_generators(self):
        cfg = AppConfig.model_validate({
            "runtime": {"concurrency": 1, "timeout": 30000},
            "browser": {"headless": True, "slow_mo": 0},
            "pages": [{"name": "static", "url": "https://example.com"}],
            "page_generators": [
                {
                    "name": "dyn",
                    "type": "ids",
                    "ids": [1, 2],
                    "template": {"name": "dyn_{{ id }}", "url": "https://example.com/{{ id }}"},
                }
            ],
        })
        assert len(cfg.pages) == 1  # static pages only before expansion
        assert len(cfg.page_generators) == 1

    def test_app_config_with_login(self):
        cfg = AppConfig.model_validate({
            "runtime": {"concurrency": 1, "timeout": 30000},
            "browser": {"headless": True, "slow_mo": 0},
            "pages": [],
            "login": {
                "enabled": True,
                "mode": "manual",
                "login_url": "https://example.com/login",
                "check": {"type": "selector", "url": "https://example.com/home", "selector": ".user"},
                "manual": {"success_selector": ".user"},
            },
        })
        assert cfg.login.enabled is True
        assert cfg.login.check.type == "selector"

    def test_backward_compatible_no_login(self):
        """Existing configs without login/page_generators must still validate."""
        cfg = AppConfig.model_validate({
            "runtime": {"concurrency": 1, "timeout": 30000},
            "browser": {"headless": True, "slow_mo": 0},
            "pages": [{"name": "test", "url": "https://example.com"}],
        })
        assert cfg.login is None
        assert cfg.page_generators == []
        assert len(cfg.pages) == 1

    def test_generated_pages_extra_fields_allowed(self):
        """Generated pages with _generated and _generator should validate (extra="allow")."""
        cfg = AppConfig.model_validate({
            "runtime": {"concurrency": 1, "timeout": 30000},
            "browser": {"headless": True, "slow_mo": 0},
            "pages": [
                {"name": "gen_page", "url": "https://example.com", "_generated": True, "_generator": "ids_gen"},
            ],
        })
        assert cfg.pages[0].name == "gen_page"


class TestUnknownGeneratorType:
    def test_skips_unknown_type(self, caplog):
        config = {
            "page_generators": [
                {
                    "name": "bad_type",
                    "type": "database",
                    "template": {"name": "x", "url": "https://x.com"},
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert pages == []

    def test_unknown_type_continues_to_next_generator(self):
        config = {
            "page_generators": [
                {
                    "name": "bad",
                    "type": "rpc",
                    "template": {"name": "x", "url": "https://x.com"},
                },
                {
                    "name": "good",
                    "type": "ids",
                    "ids": [1, 2],
                    "template": {"name": "p{{ id }}", "url": "https://x.com/{{ id }}"},
                },
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 2
        assert pages[0]["_generator"] == "good"


class TestMultipleGenerators:
    def test_two_generators(self):
        config = {
            "page_generators": [
                {
                    "name": "a_gen",
                    "type": "ids",
                    "ids": [1],
                    "template": {"name": "a", "url": "https://x.com/a"},
                },
                {
                    "name": "b_gen",
                    "type": "ids",
                    "ids": [2, 3],
                    "template": {"name": "b_{{ id }}", "url": "https://x.com/b/{{ id }}"},
                },
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 3
        assert pages[0]["_generator"] == "a_gen"
        assert pages[1]["_generator"] == "b_gen"
        assert pages[2]["_generator"] == "b_gen"

    def test_ids_and_list_mixed(self):
        config = {
            "page_generators": [
                {
                    "name": "ids_gen",
                    "type": "ids",
                    "ids": ["a"],
                    "template": {"name": "{{ id }}", "url": "https://x.com/{{ id }}"},
                },
                {
                    "name": "list_gen",
                    "type": "list",
                    "values": [{"env": "dev"}],
                    "template": {"name": "{{ env }}", "url": "https://x.com/{{ env }}"},
                },
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 2


class TestCustomIdField:
    def test_custom_id_field(self):
        config = {
            "page_generators": [
                {
                    "name": "custom",
                    "type": "ids",
                    "id_field": "resource_name",
                    "ids": ["pod-a", "pod-b"],
                    "template": {
                        "name": "{{ resource_name }}",
                        "url": "https://x.com/{{ resource_name }}",
                    },
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 2
        assert pages[0]["name"] == "pod-a"
        assert pages[1]["name"] == "pod-b"


class TestLoadAndValidateConfigIntegration:
    """Test the full load_and_validate_config pipeline with page generators."""

    def test_static_plus_generated_pages(self, tmp_path):
        """load_and_validate_config expands generators and merges with static pages."""
        import yaml
        config_file = tmp_path / "test_config.yaml"
        config = {
            "vars": {"base_url": "https://example.com"},
            "runtime": {"concurrency": 1, "timeout": 30000},
            "browser": {"headless": True, "slow_mo": 0},
            "pages": [
                {"name": "static_home", "url": "{{ base_url }}/home"},
            ],
            "page_generators": [
                {
                    "name": "ids_gen",
                    "type": "ids",
                    "ids": [100, 200],
                    "template": {"name": "item_{{ id }}", "url": "{{ base_url }}/items/{{ id }}"},
                }
            ],
        }
        config_file.write_text(yaml.dump(config))

        app_cfg, rendered = load_and_validate_config(str(config_file))
        assert len(app_cfg.pages) == 3  # 1 static + 2 generated
        assert app_cfg.pages[0].name == "static_home"
        assert app_cfg.pages[1].name == "item_100"
        assert app_cfg.pages[2].name == "item_200"
        # Rendered config has pages expanded
        assert len(rendered["pages"]) == 3

    def test_only_generated_pages(self, tmp_path):
        """Config with only page_generators and no static pages."""
        import yaml
        config_file = tmp_path / "test_config.yaml"
        config = {
            "runtime": {"concurrency": 1, "timeout": 30000},
            "browser": {"headless": True, "slow_mo": 0},
            "page_generators": [
                {
                    "name": "gen",
                    "type": "list",
                    "values": [{"name": "p1"}, {"name": "p2"}],
                    "template": {"name": "{{ name }}", "url": "https://x.com/{{ name }}"},
                }
            ],
        }
        config_file.write_text(yaml.dump(config))

        app_cfg, rendered = load_and_validate_config(str(config_file))
        assert len(app_cfg.pages) == 2
        assert app_cfg.pages[0].name == "p1"
        assert app_cfg.pages[1].name == "p2"

    def test_list_pages_includes_generated(self, tmp_path, capsys):
        """When --list-pages is used, generated pages should appear."""
        import sys
        import yaml
        config_file = tmp_path / "test_config.yaml"
        config = {
            "runtime": {"concurrency": 1, "timeout": 30000},
            "browser": {"headless": True, "slow_mo": 0},
            "pages": [{"name": "static", "url": "https://x.com/home"}],
            "page_generators": [
                {
                    "name": "gen",
                    "type": "ids",
                    "ids": [1],
                    "template": {"name": "gen_{{ id }}", "url": "https://x.com/{{ id }}"},
                }
            ],
        }
        config_file.write_text(yaml.dump(config))

        # Simulate --list-pages by directly calling load_and_validate_config
        # then iterating pages
        app_cfg, _ = load_and_validate_config(str(config_file))
        names = [p.name or p.url for p in app_cfg.pages]
        assert "static" in names
        assert "gen_1" in names
        assert len(names) == 2

    def test_invalid_generated_page_missing_url_raises(self, tmp_path):
        """A generated page template without 'url' should fail Pydantic validation."""
        import yaml
        config_file = tmp_path / "test_config.yaml"
        config = {
            "runtime": {"concurrency": 1, "timeout": 30000},
            "browser": {"headless": True, "slow_mo": 0},
            "page_generators": [
                {
                    "name": "bad",
                    "type": "ids",
                    "ids": [1],
                    "template": {"name": "no_url"},  # missing required 'url' field
                }
            ],
        }
        config_file.write_text(yaml.dump(config))

        with pytest.raises(SystemExit, match="Config validation failed"):
            load_and_validate_config(str(config_file))

    def test_vars_available_in_template(self, tmp_path):
        """Global vars should be available inside page_generator templates."""
        import yaml
        config_file = tmp_path / "test_config.yaml"
        config = {
            "vars": {"env": "prod", "region": "us-east-1"},
            "runtime": {"concurrency": 1, "timeout": 30000},
            "browser": {"headless": True, "slow_mo": 0},
            "page_generators": [
                {
                    "name": "env_gen",
                    "type": "ids",
                    "ids": ["svc-a"],
                    "template": {
                        "name": "{{ id }}-{{ env }}",
                        "url": "https://{{ region }}.example.com/{{ id }}",
                    },
                }
            ],
        }
        config_file.write_text(yaml.dump(config))

        app_cfg, _ = load_and_validate_config(str(config_file))
        assert len(app_cfg.pages) == 1
        assert app_cfg.pages[0].name == "svc-a-prod"
        assert app_cfg.pages[0].url == "https://us-east-1.example.com/svc-a"

    def test_config_with_both_login_and_generators(self, tmp_path):
        """Config with login + page_generators should validate correctly."""
        import yaml
        config_file = tmp_path / "test_config.yaml"
        config = {
            "runtime": {"concurrency": 1, "timeout": 30000},
            "browser": {"headless": True, "slow_mo": 0},
            "login": {
                "enabled": True,
                "mode": "manual",
                "login_url": "https://example.com/login",
                "manual": {"success_selector": ".user"},
            },
            "page_generators": [
                {
                    "name": "gen",
                    "type": "ids",
                    "ids": [1],
                    "template": {"name": "p{{ id }}", "url": "https://example.com/{{ id }}"},
                }
            ],
        }
        config_file.write_text(yaml.dump(config))

        app_cfg, _ = load_and_validate_config(str(config_file))
        assert app_cfg.login.enabled is True
        assert app_cfg.login.mode == "manual"
        assert len(app_cfg.pages) == 1

    def test_deep_merge_with_generators(self, tmp_path):
        """Merging two configs where second adds page_generators."""
        import yaml
        base_file = tmp_path / "base.yaml"
        base_config = {
            "vars": {"host": "x.com"},
            "runtime": {"concurrency": 1, "timeout": 30000},
            "browser": {"headless": True, "slow_mo": 0},
            "pages": [{"name": "base_page", "url": "https://x.com"}],
        }
        base_file.write_text(yaml.dump(base_config))

        override_file = tmp_path / "override.yaml"
        override_config = {
            "page_generators": [
                {
                    "name": "extra",
                    "type": "ids",
                    "ids": [10],
                    "template": {"name": "extra_{{ id }}", "url": "https://x.com/{{ id }}"},
                }
            ],
        }
        override_file.write_text(yaml.dump(override_config))

        app_cfg, _ = load_and_validate_config(str(base_file), str(override_file))
        assert len(app_cfg.pages) == 2  # 1 base + 1 generated


class TestExpandCsv:
    def test_basic_csv(self, tmp_path):
        csv_file = tmp_path / "resources.csv"
        csv_file.write_text("resource_id,system_name\n1001,HIS\n1002,LIS\n1003,EMR\n", encoding="utf-8")

        config = {
            "page_generators": [
                {
                    "name": "csv_gen",
                    "type": "csv",
                    "source": str(csv_file),
                    "template": {
                        "name": "resource_{{ resource_id }}",
                        "url": "https://example.com/ids/{{ resource_id }}/query",
                    },
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 3
        assert pages[0]["name"] == "resource_1001"
        assert pages[0]["url"] == "https://example.com/ids/1001/query"
        assert pages[0]["_generator"] == "csv_gen"
        assert pages[1]["name"] == "resource_1002"
        assert pages[2]["name"] == "resource_1003"

    def test_csv_with_global_vars(self, tmp_path):
        csv_file = tmp_path / "pods.csv"
        csv_file.write_text("pod_name,env\npod-a,prod\npod-b,staging\n", encoding="utf-8")

        config = {
            "page_generators": [
                {
                    "name": "pod_gen",
                    "type": "csv",
                    "source": str(csv_file),
                    "template": {
                        "name": "{{ pod_name }}-{{ env }}",
                        "url": "{{ base_url }}/pods/{{ pod_name }}?env={{ env }}",
                    },
                }
            ]
        }
        pages = expand_page_generators(config, {"base_url": "https://k8s.example.com"})
        assert len(pages) == 2
        assert pages[0]["name"] == "pod-a-prod"
        assert pages[0]["url"] == "https://k8s.example.com/pods/pod-a?env=prod"
        assert pages[1]["name"] == "pod-b-staging"

    def test_csv_empty_file(self, tmp_path):
        csv_file = tmp_path / "empty.csv"
        csv_file.write_text("id,name\n", encoding="utf-8")

        config = {
            "page_generators": [
                {
                    "name": "empty_gen",
                    "type": "csv",
                    "source": str(csv_file),
                    "template": {"name": "{{ id }}", "url": "https://x.com/{{ id }}"},
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert pages == []

    def test_csv_file_not_found(self, tmp_path):
        config = {
            "page_generators": [
                {
                    "name": "missing",
                    "type": "csv",
                    "source": "/nonexistent/file.csv",
                    "template": {"name": "x", "url": "https://x.com"},
                }
            ]
        }
        with pytest.raises(FileNotFoundError, match="CSV source file not found"):
            expand_page_generators(config, {})

    def test_csv_missing_source(self):
        config = {
            "page_generators": [
                {
                    "name": "no_source",
                    "type": "csv",
                    "template": {"name": "x", "url": "https://x.com"},
                }
            ]
        }
        with pytest.raises(ValueError, match="requires 'source'"):
            expand_page_generators(config, {})

    def test_csv_with_max_pages(self, tmp_path):
        csv_file = tmp_path / "many.csv"
        csv_file.write_text("id\n" + "\n".join(str(i) for i in range(10)), encoding="utf-8")

        config = {
            "page_generators": [
                {
                    "name": "limited",
                    "type": "csv",
                    "source": str(csv_file),
                    "max_pages": 3,
                    "template": {"name": "p{{ id }}", "url": "https://x.com/{{ id }}"},
                }
            ]
        }
        with pytest.raises(ValueError, match="exceeding max_pages"):
            expand_page_generators(config, {})

    def test_csv_utf8_bom(self, tmp_path):
        csv_file = tmp_path / "bom.csv"
        with open(csv_file, "w", encoding="utf-8-sig") as f:
            f.write("name,city\n张三,北京\n")

        config = {
            "page_generators": [
                {
                    "name": "bom_gen",
                    "type": "csv",
                    "source": str(csv_file),
                    "template": {"name": "{{ name }}", "url": "https://x.com/{{ city }}"},
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 1
        assert pages[0]["name"] == "张三"
        assert pages[0]["url"] == "https://x.com/北京"


class TestExpandJson:
    def test_basic_json_array(self, tmp_path):
        json_file = tmp_path / "data.json"
        json_file.write_text(json.dumps([
            {"id": "1001", "name": "HIS"},
            {"id": "1002", "name": "LIS"},
        ]), encoding="utf-8")

        config = {
            "page_generators": [
                {
                    "name": "json_gen",
                    "type": "json",
                    "source": str(json_file),
                    "template": {
                        "name": "resource_{{ id }}",
                        "url": "https://example.com/ids/{{ id }}",
                    },
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 2
        assert pages[0]["name"] == "resource_1001"
        assert pages[0]["_generator"] == "json_gen"
        assert pages[1]["name"] == "resource_1002"

    def test_json_with_items_path(self, tmp_path):
        json_file = tmp_path / "data.json"
        json_file.write_text(json.dumps({
            "status": "ok",
            "items": [
                {"id": "a", "region": "us"},
                {"id": "b", "region": "eu"},
            ],
        }), encoding="utf-8")

        config = {
            "page_generators": [
                {
                    "name": "json_path_gen",
                    "type": "json",
                    "source": str(json_file),
                    "items_path": "$.items",
                    "template": {
                        "name": "{{ id }}-{{ region }}",
                        "url": "https://example.com/{{ id }}",
                    },
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 2
        assert pages[0]["name"] == "a-us"
        assert pages[1]["name"] == "b-eu"

    def test_json_items_path_without_dollar(self, tmp_path):
        json_file = tmp_path / "data.json"
        json_file.write_text(json.dumps({"data": {"results": [{"x": 1}, {"x": 2}]}}), encoding="utf-8")

        config = {
            "page_generators": [
                {
                    "name": "nested",
                    "type": "json",
                    "source": str(json_file),
                    "items_path": "data.results",
                    "template": {"name": "item_{{ x }}", "url": "https://x.com/{{ x }}"},
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 2
        assert pages[0]["name"] == "item_1"
        assert pages[1]["name"] == "item_2"

    def test_json_scalar_array(self, tmp_path):
        json_file = tmp_path / "ids.json"
        json_file.write_text(json.dumps(["pod-a", "pod-b", "pod-c"]), encoding="utf-8")

        config = {
            "page_generators": [
                {
                    "name": "scalar_gen",
                    "type": "json",
                    "source": str(json_file),
                    "template": {
                        "name": "pod_{{ value }}",
                        "url": "https://k8s.example.com/pods/{{ value }}",
                    },
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 3
        assert pages[0]["name"] == "pod_pod-a"
        assert pages[2]["name"] == "pod_pod-c"

    def test_json_with_global_vars(self, tmp_path):
        json_file = tmp_path / "data.json"
        json_file.write_text(json.dumps([{"env": "dev"}, {"env": "prod"}]), encoding="utf-8")

        config = {
            "page_generators": [
                {
                    "name": "env_gen",
                    "type": "json",
                    "source": str(json_file),
                    "template": {
                        "name": "{{ env }}-dashboard",
                        "url": "{{ base_url }}/{{ env }}/dashboard",
                    },
                }
            ]
        }
        pages = expand_page_generators(config, {"base_url": "https://monitor.example.com"})
        assert pages[0]["url"] == "https://monitor.example.com/dev/dashboard"
        assert pages[1]["url"] == "https://monitor.example.com/prod/dashboard"

    def test_json_empty_array(self, tmp_path):
        json_file = tmp_path / "empty.json"
        json_file.write_text("[]", encoding="utf-8")

        config = {
            "page_generators": [
                {
                    "name": "empty",
                    "type": "json",
                    "source": str(json_file),
                    "template": {"name": "x", "url": "https://x.com"},
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert pages == []

    def test_json_file_not_found(self):
        config = {
            "page_generators": [
                {
                    "name": "missing",
                    "type": "json",
                    "source": "/nonexistent/data.json",
                    "template": {"name": "x", "url": "https://x.com"},
                }
            ]
        }
        with pytest.raises(FileNotFoundError, match="JSON source file not found"):
            expand_page_generators(config, {})

    def test_json_items_path_not_found(self, tmp_path):
        json_file = tmp_path / "data.json"
        json_file.write_text(json.dumps({"a": 1}), encoding="utf-8")

        config = {
            "page_generators": [
                {
                    "name": "bad_path",
                    "type": "json",
                    "source": str(json_file),
                    "items_path": "$.items",
                    "template": {"name": "x", "url": "https://x.com"},
                }
            ]
        }
        with pytest.raises(ValueError, match="items_path"):
            expand_page_generators(config, {})

    def test_json_not_a_list(self, tmp_path):
        json_file = tmp_path / "data.json"
        json_file.write_text(json.dumps({"key": "value"}), encoding="utf-8")

        config = {
            "page_generators": [
                {
                    "name": "not_list",
                    "type": "json",
                    "source": str(json_file),
                    "template": {"name": "x", "url": "https://x.com"},
                }
            ]
        }
        with pytest.raises(ValueError, match="must be a list"):
            expand_page_generators(config, {})


class TestExpandXlsx:
    def test_basic_xlsx(self, tmp_path):
        import openpyxl

        xlsx_file = tmp_path / "resources.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["resource_id", "system_name"])
        ws.append(["1001", "HIS"])
        ws.append(["1002", "LIS"])
        wb.save(str(xlsx_file))

        config = {
            "page_generators": [
                {
                    "name": "xlsx_gen",
                    "type": "xlsx",
                    "source": str(xlsx_file),
                    "template": {
                        "name": "resource_{{ resource_id }}",
                        "url": "https://example.com/ids/{{ resource_id }}",
                    },
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 2
        assert pages[0]["name"] == "resource_1001"
        assert pages[0]["_generator"] == "xlsx_gen"
        assert pages[1]["name"] == "resource_1002"

    def test_xlsx_with_global_vars(self, tmp_path):
        import openpyxl

        xlsx_file = tmp_path / "pods.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["pod_name", "namespace"])
        ws.append(["pod-a", "prod"])
        ws.append(["pod-b", "staging"])
        wb.save(str(xlsx_file))

        config = {
            "page_generators": [
                {
                    "name": "xlsx_vars",
                    "type": "xlsx",
                    "source": str(xlsx_file),
                    "template": {
                        "name": "{{ pod_name }}",
                        "url": "{{ base_url }}/{{ namespace }}/pods/{{ pod_name }}",
                    },
                }
            ]
        }
        pages = expand_page_generators(config, {"base_url": "https://k8s.example.com"})
        assert len(pages) == 2
        assert pages[0]["url"] == "https://k8s.example.com/prod/pods/pod-a"
        assert pages[1]["url"] == "https://k8s.example.com/staging/pods/pod-b"

    def test_xlsx_with_sheet_name(self, tmp_path):
        import openpyxl

        xlsx_file = tmp_path / "multi.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "ignored"
        ws1.append(["header"])
        ws1.append([1])

        ws2 = wb.create_sheet("resources")
        ws2.append(["id", "label"])
        ws2.append(["r1", "Resource 1"])
        ws2.append(["r2", "Resource 2"])
        wb.save(str(xlsx_file))

        config = {
            "page_generators": [
                {
                    "name": "sheet_gen",
                    "type": "xlsx",
                    "source": str(xlsx_file),
                    "sheet_name": "resources",
                    "template": {"name": "{{ id }}", "url": "https://x.com/{{ id }}"},
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert len(pages) == 2
        assert pages[0]["name"] == "r1"
        assert pages[1]["name"] == "r2"

    def test_xlsx_empty_file(self, tmp_path):
        import openpyxl

        xlsx_file = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id"])  # header only, no data rows
        wb.save(str(xlsx_file))

        config = {
            "page_generators": [
                {
                    "name": "empty_gen",
                    "type": "xlsx",
                    "source": str(xlsx_file),
                    "template": {"name": "{{ id }}", "url": "https://x.com/{{ id }}"},
                }
            ]
        }
        pages = expand_page_generators(config, {})
        assert pages == []

    def test_xlsx_file_not_found(self):
        config = {
            "page_generators": [
                {
                    "name": "missing",
                    "type": "xlsx",
                    "source": "/nonexistent/file.xlsx",
                    "template": {"name": "x", "url": "https://x.com"},
                }
            ]
        }
        with pytest.raises(FileNotFoundError, match="XLSX source file not found"):
            expand_page_generators(config, {})

    def test_xlsx_missing_source(self):
        config = {
            "page_generators": [
                {
                    "name": "no_source",
                    "type": "xlsx",
                    "template": {"name": "x", "url": "https://x.com"},
                }
            ]
        }
        with pytest.raises(ValueError, match="requires 'source'"):
            expand_page_generators(config, {})
